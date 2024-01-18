import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill


# First, let's re-import the Excel files and do the calculations in pandas
df1 = pd.read_excel('/mnt/data/Higginsianum_Effectors.xlsx')
df2 = pd.read_excel('/mnt/data/Destructivum_Expression.xlsx')

# Clean up 'Gene ID' by removing trailing digits after a period
df1['Gene ID'] = df1['Gene ID'].str.replace(r'\.\d*$', '', regex=True)
df2['Gene ID'] = df2['Gene ID'].str.replace(r'\.\d*$', '', regex=True)

# Drop rows where 'Gene ID' is NaN in df2
df2_cleaned = df2.dropna(subset=['Gene ID'])

# Perform a left merge
merged_df = pd.merge(df1, df2_cleaned, on='Gene ID', how='left')

# Calculate DIFF columns
merged_df['Higg DIFF'] = merged_df['BP'] - merged_df['NP']
merged_df['Destr DIFF'] = merged_df['48h'] - merged_df['72h']

# Now let's export this merged DataFrame to Excel and then apply the coloring using openpyxl
temp_output_path = '/mnt/data/Merged_Higginsianum_Destructivum_For_Coloring.xlsx'
merged_df.to_excel(temp_output_path, index=False)

# Define a function to apply the coloring logic in Excel
def apply_coloring_to_excel(file_path):
    # Load the workbook and worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Define the fill colors
    grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')

    # Apply coloring based on the conditions
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        higg_diff = row[-2].value  # Second to last column is Higg DIFF
        destr_diff = row[-1].value # Last column is Destr DIFF

        # Determine the fill color
        if destr_diff is None:  # Missing Destructivum Gene ID
            fill_color = grey_fill
        elif (higg_diff > 0 and destr_diff > 0) or (higg_diff < 0 and destr_diff < 0):
            fill_color = green_fill
        else:
            fill_color = red_fill

        # Apply the fill color to the row
        for cell in row:
            cell.fill = fill_color

    # Save the colored workbook
    colored_output_path = file_path.replace('.xlsx', '_Colored.xlsx')
    wb.save(colored_output_path)
    return colored_output_path

# Apply coloring to the Excel file
final_colored_file_path = apply_coloring_to_excel(temp_output_path)
final_colored_file_path